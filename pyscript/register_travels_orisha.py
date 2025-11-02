@service
def register_orisha_allowance(day=1, month=1, year=2025, from_travel="", to_travel="", home_allowance=0, travelled_km=0, logline=2):
    """yaml
name: Register Orisha allowance
description: hello_world service example using pyscript.
fields:
    day:
        description: Day of the month of the allowance
        example: 1
        required: true
        selector:
            number:
                min:1
                max:31
    month:
        description: Month number of the allowance
        example: 1
        required: true
        selector:
            number:
                min:1
                max:12
    year:
        description: Year of the allowance
        example: 1
        required: true
        selector:
            number:
                min:2025
                max:2100
    from_travel:
        description: Travelled from
        example: Home
        required: true
        selector:
            text:
    to_travel:
        description: Travelled to
        example: Office Zwolle
        required: true
        selector:
            text:
    home_allowance:
        description: Working from home?
        example: 1
        required: true
        selector:
            number:
                min:0
                max:1
    travelled_km:
        description: Number of km for the allowance
        example: 1
        required: true
        selector:
            number:
                min:0
                max:500
    logline:
        description: logline to put it on
        example: 1
        required: true
        selector:
            number:
                min:6
                max:37
"""
    from openpyxl import load_workbook

    # Format date
    month_str = f"{day}-{month}"
    date_str = f"{month_str}-{year}"

    # File path
    xltx_path = f"/config/www/orisha/offereins_travels_{year}_{month:02d}.xltx"

    # Load workbook and worksheet
    wb = load_workbook(filename=xltx_path)
    ws = wb.active

    # Insert values into the specified row
    ws.cell(row=logline, column=1, value=date_str)
    ws.cell(row=logline, column=2, value=from_travel)
    ws.cell(row=logline, column=3, value=to_travel)
    ws.cell(row=logline, column=4, value=home_allowance)
    ws.cell(row=logline, column=5, value=travelled_km)

    # Save workbook
    wb.template = True
    wb.save(xltx_path)
